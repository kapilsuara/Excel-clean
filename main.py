import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.utils import range_boundaries
import numpy as np
import re
import logging
import os
import tempfile
import shutil
import json
from dotenv import load_dotenv
import datetime

# Load environment variables
load_dotenv()

# Set Streamlit to wide layout
st.set_page_config(layout="wide")

# Set up logging
logging.basicConfig(filename='data_cleaning_log.txt', level=logging.INFO, 
                    format='%(asctime)s - %(levelname)s - %(message)s')

# Get Anthropic API key from environment
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY")
if not ANTHROPIC_API_KEY:
    st.error("Anthropic API key not found in .env file. Please add ANTHROPIC_API_KEY to your .env file.")
    st.stop()

# Simple CSS for layout
st.markdown("""
<style>
.main .block-container {
    max-width: 100%;
    padding-left: 1rem;
    padding-right: 1rem;
}
</style>
""", unsafe_allow_html=True)

def repair_corrupted_excel(input_path, temp_path='repaired.xlsx'):
    """Repair Excel file and return workbook"""
    try:
        if not os.access(input_path, os.W_OK):
            shutil.copy2(input_path, temp_path)
            os.chmod(temp_path, 0o666)
            return openpyxl.load_workbook(temp_path), temp_path
        else:
            return openpyxl.load_workbook(input_path), input_path
    except Exception as e:
        logging.error(f"Error opening Excel file: {e}")
        try:
            df = pd.read_excel(input_path, engine='openpyxl', dtype_backend='numpy_nullable')
            df.to_excel(temp_path, index=False, engine='openpyxl')
            return openpyxl.load_workbook(temp_path), temp_path
        except Exception as e2:
            logging.error(f"Failed to repair Excel file: {e2}")
            raise

def fill_merged_cells(ws):
    """Fill merged cells with the value from the top-left cell"""
    for merged_range in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        top_left_value = ws.cell(row=min_row, column=min_col).value
        ws.unmerge_cells(str(merged_range))
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row=row, column=col, value=top_left_value)

def remove_hidden_rows_columns(ws):
    """Remove hidden rows and columns from worksheet"""
    visible_rows = []
    for row in range(1, ws.max_row + 1):
        if not ws.row_dimensions[row].hidden:
            visible_rows.append(row)
    
    visible_cols = []
    for col in range(1, ws.max_column + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        if not ws.column_dimensions[col_letter].hidden:
            visible_cols.append(col)
    
    if len(visible_rows) == ws.max_row and len(visible_cols) == ws.max_column:
        return
    
    new_ws = ws._parent.create_sheet(title=ws.title + "_visible")
    for i, row in enumerate(visible_rows, start=1):
        for j, col in enumerate(visible_cols, start=1):
            new_ws.cell(row=i, column=j).value = ws.cell(row=row, column=col).value
    
    ws._parent.remove(ws)
    new_ws.title = ws.title

def detect_header_row(temp_path, sheet_name):
    """Use AI to detect the header row index"""
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
        
        wb = openpyxl.load_workbook(temp_path, read_only=True)
        ws = wb[sheet_name]
        
        rows = []
        for row in ws.iter_rows(max_row=20):
            row_values = [cell.value for cell in row if cell.value is not None]
            rows.append(row_values)
        
        wb.close()
        
        prompt = f"""Here are the first rows of an Excel sheet as a list of lists (empty cells omitted):
{json.dumps(rows, ensure_ascii=False)}

Identify the 0-based index of the row that most likely contains the column headers. Look for a row with strings that resemble labels or field names (e.g., 'Name', 'Date', 'Total', 'Health ATM'), not data values. Consider context: headers often have descriptive words, while data rows may have numbers or specific data entries. If no clear header row is found, or if data starts immediately, return -1.
Output only the integer index or -1, nothing else."""
        
        response = client.messages.create(
            model="claude-3-5-sonnet-20241022",
            max_tokens=10,
            messages=[{"role": "user", "content": prompt}]
        )
        
        header_index_str = response.content[0].text.strip()
        try:
            header_index = int(header_index_str)
            return header_index
        except:
            return -1
        
    except ImportError:
        logging.error("Anthropic library not installed")
        return -1
    except Exception as e:
        logging.error(f"Error detecting header: {e}")
        return -1

def infer_data_types(df, changes_log):
    """Infer and correct data types for columns without losing precision"""
    for col in df.columns:
        if df[col].dtype.kind in 'O':  # object
            # Try integer
            try:
                df_temp = df[col].astype('Int64')
                if not df_temp.isna().all():
                    df[col] = df_temp
                    changes_log.append(f"✓ Converted '{col}' to Int64")
            except:
                # Try numeric (float)
                try:
                    df_temp = pd.to_numeric(df[col], errors='raise')
                    if not pd.isna(df_temp).all():
                        df[col] = df_temp
                        changes_log.append(f"✓ Converted '{col}' to numeric")
                except:
                    # Try datetime
                    try:
                        df_temp = pd.to_datetime(df[col], errors='coerce')
                        if not df_temp.isna().all():
                            df[col] = df_temp
                            changes_log.append(f"✓ Converted '{col}' to datetime")
                    except:
                        pass
            # If kept as object due to large values
            if df[col].dtype.kind in 'O':
                values = df[col].dropna().astype(str)
                if all(re.match(r'^-?\d+$', v) for v in values):
                    changes_log.append(f"⚠️ Kept '{col}' as string: large integers potentially out of range")

def standardize_dates(df, changes_log):
    """Standardize date formats"""
    date_cols = df.select_dtypes(include=['datetime64']).columns
    for col in date_cols:
        df[col] = df[col].dt.strftime('%Y-%m-%d')
    if len(date_cols) > 0:
        changes_log.append(f"✓ Standardized formats in {len(date_cols)} date columns")

def validate_ranges(df, changes_log):
    """Validate data ranges and flag issues"""
    numeric_cols = df.select_dtypes(include=[np.number, 'Int64']).columns
    for col in numeric_cols:
        non_null = df[col].dropna()
        if len(non_null) == 0:
            continue
        if (non_null < 0).any():
            changes_log.append(f"⚠️ Negative values found in '{col}' - review if appropriate")
        mean = non_null.mean()
        std = non_null.std()
        outliers = ((non_null - mean).abs() > 3 * std).sum()
        if outliers > 0:
            changes_log.append(f"⚠️ {outliers} potential outliers in '{col}'")

def clean_excel_basic(input_path, output_path, sheet_name=None):
    """Basic Excel cleaning with classical Python"""
    changes_log = []
    
    try:
        # Load original for initial info
        original_df = pd.read_excel(input_path, sheet_name=sheet_name if sheet_name else 0, dtype_backend='numpy_nullable')
        changes_log.append(f"✓ Loaded dataset: {original_df.shape[0]} rows × {original_df.shape[1]} columns")
    except Exception as e:
        changes_log.append(f"❌ Error loading data: {str(e)}")
        return pd.DataFrame(), changes_log
    
    try:
        # Repair Excel file
        wb, writable_path = repair_corrupted_excel(input_path, temp_path='temp_writable.xlsx')
        
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active
            sheet_name = ws.title
        
        # Handle extra sheets
        if len(wb.sheetnames) > 1:
            changes_log.append(f"⚠️ Multiple sheets found: {wb.sheetnames}. Processing only '{sheet_name}'")
        
        # Handle merged cells
        merged_count = len(list(ws.merged_cells.ranges))
        if merged_count > 0:
            fill_merged_cells(ws)
            changes_log.append(f"✓ Processed {merged_count} merged cell ranges")
        
        # Remove hidden rows/columns
        remove_hidden_rows_columns(ws)
        changes_log.append("✓ Removed hidden rows and columns")
        
        # Save to temp file
        temp_path = 'temp_clean.xlsx'
        wb.save(temp_path)
        
        # Detect header row with AI
        header_index = detect_header_row(temp_path, sheet_name)
        changes_log.append(f"AI detected header row at index {header_index}")
        
        # Load into pandas
        if header_index == -1:
            df = pd.read_excel(temp_path, sheet_name=sheet_name, header=None, dtype_backend='numpy_nullable')
            changes_log.append("✓ Loaded without header row")
        else:
            df = pd.read_excel(temp_path, sheet_name=sheet_name, header=header_index, dtype_backend='numpy_nullable')
            changes_log.append(f"✓ Loaded with header at row {header_index + 1}")
        
        # Clean column names and ensure uniqueness
        if len(df.columns) == 0:
            changes_log.append("⚠️ No columns found in the dataset")
            return pd.DataFrame(), changes_log
        
        # Handle column names
        new_columns = []
        seen_names = {}
        for col in df.columns:
            if pd.isna(col) or col == '':
                base_name = f"Unnamed_{len(new_columns)}"
            else:
                col_str = str(col).strip().title().replace(' ', '_')
                col_str = re.sub(r'[^\w]', '_', col_str)
                base_name = col_str if col_str else f"Unnamed_{len(new_columns)}"
            
            # Ensure unique column names
            if base_name in seen_names:
                seen_names[base_name] += 1
                new_name = f"{base_name}_{seen_names[base_name]}"
            else:
                seen_names[base_name] = 0
                new_name = base_name
            new_columns.append(new_name)
        
        df.columns = new_columns
        changes_log.append(f"✓ Standardized {len(df.columns)} column names and ensured uniqueness")
        if any(v > 0 for v in seen_names.values()):
            changes_log.append(f"⚠️ Resolved duplicate column names by appending suffixes: {seen_names}")
        
        # Remove empty rows and completely empty unnamed columns
        initial_shape = df.shape
        df = df.dropna(how='all', axis=0).reset_index(drop=True)
        empty_cols = [col for col in df.columns if col.startswith('Unnamed_') and df[col].isna().all()]
        df = df.drop(columns=empty_cols)
        
        if df.shape != initial_shape:
            changes_log.append(f"✓ Removed {initial_shape[0] - df.shape[0]} empty rows and {len(empty_cols)} empty columns")
        
        # Trim whitespace in text
        text_cols = df.select_dtypes(include=['object']).columns
        for col in text_cols:
            df[col] = df[col].astype(str).str.strip().str.replace(r'\s+', ' ', regex=True)
        
        if len(text_cols) > 0:
            changes_log.append(f"✓ Cleaned text in {len(text_cols)} columns")
        
        # Set missing values to NaN
        for col in df.columns:
            df[col] = df[col].replace('', np.nan).replace('null', np.nan).replace('NaN', np.nan)
        changes_log.append("✓ Set missing/empty values to NaN")
        
        # Infer data types
        infer_data_types(df, changes_log)
        
        # Standardize dates
        standardize_dates(df, changes_log)
        
        # Remove duplicates
        initial_len = len(df)
        df = df.drop_duplicates()
        duplicates_removed = initial_len - len(df)
        if duplicates_removed > 0:
            changes_log.append(f"✓ Removed {duplicates_removed} duplicate rows")
        
        # Validate ranges
        validate_ranges(df, changes_log)
        
        # Handle inconsistent delimiters (replace semicolons with commas in text)
        for col in text_cols:
            if col in df.columns and df[col].str.contains(';', na=False).any():
                df[col] = df[col].str.replace(';', ',')
                changes_log.append(f"✓ Standardized delimiters in '{col}'")
        
        # Encoding: ensure UTF-8
        for col in text_cols:
            if col in df.columns:
                df[col] = df[col].str.encode('utf-8', 'replace').str.decode('utf-8')
        changes_log.append("✓ Ensured UTF-8 encoding")
        
        # Save cleaned file
        df.to_excel(output_path, index=False, engine='openpyxl')
        changes_log.append(f"✅ Saved cleaned data to {output_path}")
        
        # Clean up temp files
        for temp_file in ['temp_clean.xlsx', 'temp_writable.xlsx']:
            if os.path.exists(temp_file):
                try:
                    os.remove(temp_file)
                except:
                    pass
        
        changes_log.append(f"📊 Final dataset: {df.shape[0]} rows × {df.shape[1]} columns")
        return df, changes_log
        
    except Exception as e:
        changes_log.append(f"❌ Error processing file: {str(e)}")
        return pd.DataFrame(), changes_log

def ai_analyze_df(df, changes_log):
    """Use AI to analyze metadata and provide suggestions"""
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
        
        # Get DataFrame info
        df_info = f"Columns: {list(df.columns)}\nDtypes: {df.dtypes.to_string()}\nSample data (first 5 rows):\n{df.head(5).to_string()}\nDescribe:\n{df.describe().to_string()}"
        
        prompt = f"""Analyze this DataFrame:
{df_info}

Extract metadata for each column: dtype, missing_count, missing_pct (as percentage), unique_count, sample_values (list of up to 3 unique non-null values if available, else empty list).

Analyze metadata and data for:
- Incorrect headers (e.g., first row is data, topic names). Suggest better names based on content (e.g., 'Invoice_ID' for invoice numbers).
- Poor structure, blank rows/columns (only remove completely empty columns with no name), whitespace, nulls (as NaN), data types, duplicates, invalid ranges, delimiters, encoding, date formats, extra sheets (already handled), structural mismatches, irregularities.
- Do not suggest filling missing values unless specified in context. Do not suggest removing columns unless completely empty with no name. Suggest unit standardizations (e.g., % to decimal) as optional.

Provide suggestions as actionable fixes (e.g., "Rename column 'Col1' to 'Price' because it contains currency values").

Output in JSON:
{{
  "metadata": {{ "Column1": {{ "dtype": "str", "missing_count": 0, "missing_pct": 0.0, "unique_count": 10, "sample_values": ["a", "b", "c"] }}, ... }},
  "analysis": ["Observation1", ...],
  "suggestions": ["Suggestion1", ...]
}}
"""
        
        response = client.messages.create(
            model="claude-3-5-sonnet-20241022",
            max_tokens=2000,
            messages=[{"role": "user", "content": prompt}]
        )
        
        json_str = response.content[0].text.strip()
        analysis = json.loads(json_str)
        
        changes_log.append("✓ AI analysis completed")
        return analysis, changes_log
        
    except ImportError:
        changes_log.append("❌ Anthropic library not installed. Run: pip install anthropic")
        return None, changes_log
    except json.JSONDecodeError:
        changes_log.append("❌ AI response not valid JSON")
        return None, changes_log
    except Exception as e:
        error_msg = str(e)
        if "authentication" in error_msg.lower() or "401" in error_msg:
            changes_log.append("❌ API Key Error: The Anthropic API key appears to be invalid or expired. Please update in .env")
        elif "rate_limit" in error_msg.lower():
            changes_log.append("❌ Rate limit exceeded. Please try again in a moment.")
        elif "insufficient_quota" in error_msg.lower():
            changes_log.append("❌ Insufficient quota. Please add credits to your Anthropic account.")
        else:
            changes_log.append(f"❌ AI Error: {error_msg}")
        return None, changes_log

def apply_ai_suggestions(df, selected_suggestions):
    """Apply selected AI suggestions using code generation"""
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
        
        df_info = f"Columns: {list(df.columns)}\nSample data:\n{df.head(3).to_string()}"
        
        prompt = f"""Given this DataFrame:
{df_info}

Implement these suggestions:
{json.dumps(selected_suggestions)}

Generate Python pandas code to modify 'df'. Do not fill missing values unless specified. Do not remove columns unless completely empty with no name. Only return executable code.
"""
        
        response = client.messages.create(
            model="claude-3-5-sonnet-20241022",
            max_tokens=1000,
            messages=[{"role": "user", "content": prompt}]
        )
        
        code = response.content[0].text.strip()
        
        # Execute the code safely
        local_vars = {'df': df.copy(), 'pd': pd, 'np': np, 'datetime': datetime}
        exec(code, {}, local_vars)
        
        return local_vars['df'], f"✅ Applied suggestions: {code}"
        
    except ImportError:
        return df, "❌ Anthropic library not installed. Run: pip install anthropic"
    except Exception as e:
        error_msg = str(e)
        return df, f"❌ Error applying suggestions: {error_msg}"

def apply_user_query_to_df(df, query):
    """Apply user's natural language query to DataFrame using Claude"""
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
        
        df_info = f"Columns: {list(df.columns)}\nSample data:\n{df.head(3).to_string()}"
        
        prompt = f"""Given this DataFrame:
{df_info}

User query: {query}

Generate Python pandas code to modify 'df'. Do not fill missing values. Do not remove columns unless completely empty with no name. Only return executable code.
Example: df = df.rename(columns={{'Col1': 'Name'}})
"""
        
        response = client.messages.create(
            model="claude-3-5-sonnet-20241022",
            max_tokens=300,
            messages=[{"role": "user", "content": prompt}]
        )
        
        code = response.content[0].text.strip()
        
        # Execute the code safely
        local_vars = {'df': df.copy(), 'pd': pd, 'np': np}
        exec(code, {}, local_vars)
        
        return local_vars['df'], f"✅ Applied: {code}"
        
    except ImportError:
        return df, "❌ Anthropic library not installed. Run: pip install anthropic"
    except Exception as e:
        error_msg = str(e)
        if "authentication" in error_msg.lower() or "401" in error_msg:
            return df, "❌ API Key Error: The Anthropic API key appears to be invalid or expired. Please update in .env"
        elif "rate_limit" in error_msg.lower():
            return df, "❌ Rate limit exceeded. Please try again in a moment."
        elif "insufficient_quota" in error_msg.lower():
            return df, "❌ Insufficient quota. Please add credits to your Anthropic account."
        else:
            return df, f"❌ Error: {error_msg}"

# Streamlit App
st.title("🧹 Basic Excel Data Cleaner")
st.markdown("Upload your Excel file for basic cleaning and AI-enhanced analysis")

# File uploader
uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx", "xls"])

if uploaded_file is not None:
    # Save uploaded file
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as input_tmp:
        input_tmp.write(uploaded_file.getvalue())
        input_path = input_tmp.name
    
    # Get sheet names
    try:
        xl_file = pd.ExcelFile(input_path, engine='openpyxl')
        sheet_names = xl_file.sheet_names
        
        # Sheet selection
        if len(sheet_names) > 1:
            selected_sheet = st.selectbox("Select sheet to process", sheet_names)
        else:
            selected_sheet = sheet_names[0]
            st.info(f"Processing sheet: {selected_sheet}")
        
        # Process button
        if st.button("🚀 Clean Data", type="primary"):
            with st.spinner("Processing your file..."):
                try:
                    output_path = f"cleaned_{uploaded_file.name}"
                    cleaned_df, changes = clean_excel_basic(input_path, output_path, selected_sheet)
                    
                    if not cleaned_df.empty:
                        # AI analysis
                        ai_analysis, changes = ai_analyze_df(cleaned_df, changes)
                        st.session_state.ai_analysis = ai_analysis
                        
                        # Save after basic
                        cleaned_df.to_excel(output_path, index=False, engine='openpyxl')
                        changes.append("✅ Updated cleaned file after basic cleaning")
                    
                    # Store in session state
                    st.session_state.cleaned_df = cleaned_df
                    st.session_state.changes_log = changes
                    
                    if not cleaned_df.empty:
                        st.success("✅ File processed successfully!")
                    else:
                        st.error("❌ Processing failed - check the error messages above")
                    
                except Exception as e:
                    st.error(f"Error processing file: {str(e)}")
                finally:
                    if os.path.exists(input_path):
                        os.unlink(input_path)
    
    except Exception as e:
        st.error(f"Error reading Excel file: {str(e)}")
        if os.path.exists(input_path):
            os.unlink(input_path)

# Display results if available
if 'cleaned_df' in st.session_state and not st.session_state.cleaned_df.empty:
    # Two-column layout
    left_col, right_col = st.columns([3, 1])
    
    with left_col:
        # Changes log
        st.header("📋 Cleaning Report")
        with st.expander("View Changes", expanded=True):
            for change in st.session_state.changes_log:
                if change.startswith("✅"):
                    st.success(change)
                elif change.startswith("⚠️"):
                    st.warning(change)
                elif change.startswith("❌"):
                    st.error(change)
                else:
                    st.info(change)
        
        # AI Metadata and Analysis
        if 'ai_analysis' in st.session_state and st.session_state.ai_analysis:
            st.header("🤖 AI Analysis")
            
            # Metadata
            st.subheader("Metadata")
            metadata_df = pd.DataFrame(st.session_state.ai_analysis['metadata']).T
            # Convert sample_values to string to avoid pyarrow error
            if 'sample_values' in metadata_df.columns:
                metadata_df['sample_values'] = metadata_df['sample_values'].apply(str)
            st.dataframe(metadata_df, use_container_width=True)
            
            # Analysis
            st.subheader("Observations")
            for obs in st.session_state.ai_analysis['analysis']:
                st.info(obs)
            
            # Suggestions
            st.subheader("Suggested Fixes")
            suggestions = st.session_state.ai_analysis['suggestions']
            if suggestions:
                selected = st.multiselect("Select suggestions to apply", suggestions, default=suggestions)
                if st.button("Apply Selected Suggestions"):
                    with st.spinner("Applying suggestions..."):
                        new_df, result = apply_ai_suggestions(st.session_state.cleaned_df, selected)
                        if "✅" in result:
                            st.session_state.cleaned_df = new_df
                            st.session_state.changes_log.append(result)
                            # Update output file
                            output_path = f"cleaned_{uploaded_file.name}"
                            new_df.to_excel(output_path, index=False, engine='openpyxl')
                            st.success("Suggestions applied!")
                            st.rerun()
                        else:
                            st.error(result)
            else:
                st.info("No suggestions from AI")
        
        # Data preview
        st.header("👁️ Cleaned Data")
        st.dataframe(st.session_state.cleaned_df, use_container_width=True)
        
        # Data summary
        st.subheader("📊 Data Summary")
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("Rows", st.session_state.cleaned_df.shape[0])
        with col2:
            st.metric("Columns", st.session_state.cleaned_df.shape[1])
        with col3:
            st.metric("Total Cells", st.session_state.cleaned_df.shape[0] * st.session_state.cleaned_df.shape[1])
        with col4:
            missing_values = st.session_state.cleaned_df.isna().sum().sum()
            st.metric("Missing Values", missing_values)
        
        # Download section
        st.header("💾 Download Cleaned Data")
        
        # Create download buffer
        from io import BytesIO
        output_buffer = BytesIO()
        with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
            st.session_state.cleaned_df.to_excel(writer, index=False, sheet_name='Cleaned_Data')
        
        st.download_button(
            label="📥 Download Cleaned Excel File",
            data=output_buffer.getvalue(),
            file_name=f"cleaned_{uploaded_file.name}",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    with right_col:
        # Chatbot
        st.header("💬 Data Chatbot")
        st.caption("Ask me to modify your data using natural language!")
        
        # Initialize chat history
        if 'chat_history' not in st.session_state:
            st.session_state.chat_history = []
        
        # Display chat history
        chat_container = st.container()
        with chat_container:
            for message in st.session_state.chat_history:
                if message["role"] == "user":
                    st.markdown(f"**🧑 You:** {message['content']}")
                else:
                    st.markdown(f"**🤖 Bot:** {message['content']}")
        
        # Chat input
        with st.form("chat_form", clear_on_submit=True):
            user_query = st.text_input("Your request:", placeholder="e.g., 'rename column Col1 to Name'")
            send_button = st.form_submit_button("Send")
        
        if send_button and user_query:
            # Add user message
            st.session_state.chat_history.append({"role": "user", "content": user_query})
            
            # Process query
            with st.spinner("Processing your request..."):
                new_df, result = apply_user_query_to_df(st.session_state.cleaned_df, user_query)
                
                # Update dataframe if successful
                if "✅" in result:
                    st.session_state.cleaned_df = new_df
                
                # Add bot response
                st.session_state.chat_history.append({"role": "bot", "content": result})
                
                # Rerun to show updates
                st.rerun()
        
        # Clear chat button
        if st.button("🗑️ Clear Chat History"):
            st.session_state.chat_history = []
            st.rerun()

# Information section
if 'cleaned_df' not in st.session_state:
    st.markdown("""
    ### 🔧 Features:
    - **Basic Cleaning**: Repairs files, AI-detects headers, standardizes names (ensures unique names), removes duplicates/empty rows and completely empty unnamed columns, validates ranges, handles merged/hidden cells, delimiters, encoding.
    - **AI Analysis**: Provides metadata (dtype, missing count/percent, unique count, sample values), observations, and suggestions for specified issues. No value filling or column removal unless empty.
    - **Apply Suggestions**: Select and apply AI-suggested fixes (e.g., rename columns, standardize units).
    - **Data Processing**: Sets missing to NaN, trims whitespace.
    - **AI Chatbot**: Natural language data manipulation (requires Anthropic API key in .env)
    - **Download**: Get your cleaned data as Excel file
    
    ### 📝 How to use:
    1. Upload your Excel file using the file uploader
    2. Select a sheet (if multiple sheets exist)
    3. Click "Clean Data" for basic cleaning and AI analysis
    4. Review the report, AI analysis, select suggestions to apply
    5. Use the chatbot for additional modifications
    """)