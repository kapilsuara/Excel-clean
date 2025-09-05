"""
Excel processing and cleaning functions
"""
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils import range_boundaries
import re
import logging
import tempfile
import shutil
import os
from config import TEMP_DIR

def check_and_make_writable(input_path, temp_path):
    """
    Check if file is writable, if not create a writable copy
    """
    if os.access(input_path, os.W_OK):
        return input_path
    else:
        shutil.copy2(input_path, temp_path)
        os.chmod(temp_path, 0o666)
        return temp_path

def repair_corrupted_excel(input_path, temp_path='repaired.xlsx'):
    """
    Attempt to repair corrupted Excel file and return workbook and path
    """
    try:
        wb = openpyxl.load_workbook(input_path, data_only=False, keep_vba=False)
        writable_path = check_and_make_writable(input_path, temp_path)
        return wb, writable_path
    except PermissionError:
        temp_copy = check_and_make_writable(input_path, temp_path)
        wb = openpyxl.load_workbook(temp_copy, data_only=False, keep_vba=False)
        return wb, temp_copy
    except Exception as e:
        logging.error(f"Error opening Excel file: {e}")
        try:
            df = pd.read_excel(input_path, engine='openpyxl')
            df.to_excel(temp_path, index=False)
            wb = openpyxl.load_workbook(temp_path)
            return wb, temp_path
        except Exception as e2:
            logging.error(f"Failed to repair Excel file: {e2}")
            raise

def fill_merged_cells(ws):
    """
    Fill merged cells with the value from the top-left cell
    """
    for merged_range in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        top_left_value = ws.cell(row=min_row, column=min_col).value
        
        ws.unmerge_cells(str(merged_range))
        
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row=row, column=col, value=top_left_value)

def handle_hidden_rows_columns(ws):
    """
    Unhide all hidden rows and columns
    """
    for row in ws.row_dimensions:
        ws.row_dimensions[row].hidden = False
    for col in ws.column_dimensions:
        ws.column_dimensions[col].hidden = False

def clean_text_column(series):
    """
    Clean text columns - remove special characters, normalize whitespace
    """
    if series.dtype == 'object':
        # Remove special characters but keep basic punctuation
        cleaned = series.astype(str).str.replace(r'[^\w\s\-.,;:()/]', ' ', regex=True)
        # Normalize whitespace
        cleaned = cleaned.str.replace(r'\s+', ' ', regex=True).str.strip()
        # Replace 'nan' strings with actual NaN
        cleaned = cleaned.replace('nan', np.nan)
        return cleaned
    return series

def detect_and_convert_dates(series):
    """
    Detect and convert date columns
    """
    if series.dtype == 'object':
        # Try to convert to datetime
        try:
            converted = pd.to_datetime(series, errors='coerce')
            # If more than 30% converted successfully, use the converted version
            if converted.notna().sum() / len(series) > 0.3:
                return converted
        except:
            pass
    return series

def clean_numeric_column(series):
    """
    Clean numeric columns - remove currency symbols, convert to numeric
    """
    if series.dtype == 'object':
        # Check if column contains numeric-like values
        sample = series.dropna().astype(str).head(10)
        if any(sample.str.contains(r'[\d,.$€£¥₹]', regex=True)):
            # Remove currency symbols and convert
            cleaned = series.astype(str).str.replace(r'[$€£¥₹,]', '', regex=True)
            cleaned = pd.to_numeric(cleaned, errors='coerce')
            
            # If more than 50% converted successfully, use the cleaned version
            if cleaned.notna().sum() / len(series) > 0.5:
                return cleaned
    return series

def detect_outliers(df):
    """
    Detect outliers in numeric columns using IQR method
    """
    outliers = {}
    for col in df.select_dtypes(include=[np.number]).columns:
        Q1 = df[col].quantile(0.25)
        Q3 = df[col].quantile(0.75)
        IQR = Q3 - Q1
        lower_bound = Q1 - 1.5 * IQR
        upper_bound = Q3 + 1.5 * IQR
        
        outlier_mask = (df[col] < lower_bound) | (df[col] > upper_bound)
        outlier_count = outlier_mask.sum()
        
        if outlier_count > 0:
            outliers[col] = {
                'count': outlier_count,
                'percentage': (outlier_count / len(df)) * 100,
                'lower_bound': lower_bound,
                'upper_bound': upper_bound
            }
    
    return outliers

def standardize_column_names(columns):
    """
    Standardize column names
    """
    new_columns = []
    for col in columns:
        if pd.isna(col):
            new_col = f"Unnamed_{len(new_columns)}"
        else:
            # Convert to string and clean
            new_col = str(col).strip()
            # Remove special characters
            new_col = re.sub(r'[^\w\s]', '_', new_col)
            # Replace spaces with underscores
            new_col = re.sub(r'\s+', '_', new_col)
            # Remove consecutive underscores
            new_col = re.sub(r'_+', '_', new_col)
            # Remove leading/trailing underscores
            new_col = new_col.strip('_')
            # Convert to title case
            new_col = new_col.title()
        
        # Ensure uniqueness
        if new_col in new_columns:
            counter = 1
            while f"{new_col}_{counter}" in new_columns:
                counter += 1
            new_col = f"{new_col}_{counter}"
        
        new_columns.append(new_col)
    
    return new_columns

def remove_duplicate_rows(df):
    """
    Remove duplicate rows and return cleaned DataFrame with count
    """
    initial_rows = len(df)
    df_cleaned = df.drop_duplicates()
    removed_rows = initial_rows - len(df_cleaned)
    return df_cleaned, removed_rows

def handle_missing_values(df, strategy='auto'):
    """
    Handle missing values based on strategy
    """
    changes = []
    
    for col in df.columns:
        missing_count = df[col].isna().sum()
        if missing_count > 0:
            missing_pct = (missing_count / len(df)) * 100
            
            if strategy == 'auto':
                # Auto strategy based on data type and missing percentage
                if missing_pct > 50:
                    # Too many missing values, consider dropping column
                    changes.append(f"Column '{col}' has {missing_pct:.1f}% missing values")
                elif pd.api.types.is_numeric_dtype(df[col]):
                    # Fill numeric columns with median
                    median_val = df[col].median()
                    df[col].fillna(median_val, inplace=True)
                    changes.append(f"Filled {missing_count} missing values in '{col}' with median ({median_val:.2f})")
                else:
                    # Fill text columns with empty string
                    df[col].fillna('', inplace=True)
                    changes.append(f"Filled {missing_count} missing values in '{col}' with empty string")
    
    return df, changes